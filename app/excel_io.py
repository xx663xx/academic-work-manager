from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import DB_PATH, get_assignments_for_export, get_connection, init_db


STUDENT_COLUMNS = ["ФИО", "Группа", "Курс", "Логин", "Контакт"]
TEACHER_COLUMNS = [
    "ФИО",
    "Должность",
    "Ученая степень",
    "Ученое звание",
    "Направление",
    "Контакт",
]
STATUS_LABELS = {
    "free": "свободно",
    "pending": "ожидает подтверждения",
    "confirmed": "подтверждено",
    "rejected": "отказано",
    "changed": "изменено",
}
EXPORT_COLUMN_WIDTHS = {
    "A": 28,
    "B": 12,
    "C": 8,
    "D": 18,
    "E": 42,
    "F": 28,
    "G": 24,
    "H": 20,
}


def read_rows(file_path, required_columns):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = [clean_text(cell.value) for cell in sheet[2]]
    missing_columns = [column for column in required_columns if column not in headers]

    if missing_columns:
        joined_columns = ", ".join(missing_columns)
        raise ValueError(f"Нет обязательных колонок: {joined_columns}")

    rows = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        row_data = dict(zip(headers, row))
        rows.append({column: row_data.get(column) for column in required_columns})

    if not rows:
        raise ValueError("Файл не содержит строк с данными")

    return rows


def import_students_from_excel(file_path, db_path=DB_PATH):
    init_db(db_path)
    rows = read_rows(Path(file_path), STUDENT_COLUMNS)
    validate_students(rows)

    with get_connection(db_path) as connection:
        for row in rows:
            connection.execute(
                """
                INSERT INTO students (full_name, study_group, course, login, contact)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(full_name, study_group) DO UPDATE SET
                    course = excluded.course,
                    login = excluded.login,
                    contact = excluded.contact
                """,
                (
                    clean_text(row["ФИО"]),
                    clean_text(row["Группа"]),
                    normalize_course(row["Курс"]),
                    clean_text(row["Логин"]),
                    clean_text(row["Контакт"]),
                ),
            )

    return len(rows)


def import_teachers_from_excel(file_path, db_path=DB_PATH):
    init_db(db_path)
    rows = read_rows(Path(file_path), TEACHER_COLUMNS)
    validate_teachers(rows)

    with get_connection(db_path) as connection:
        for row in rows:
            connection.execute(
                """
                INSERT INTO teachers (
                    full_name,
                    position,
                    academic_degree,
                    academic_title,
                    specialization,
                    contact
                )
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(full_name) DO UPDATE SET
                    position = excluded.position,
                    academic_degree = excluded.academic_degree,
                    academic_title = excluded.academic_title,
                    specialization = excluded.specialization,
                    contact = excluded.contact
                """,
                (
                    clean_text(row["ФИО"]),
                    clean_text(row["Должность"]),
                    clean_text(row["Ученая степень"]),
                    clean_text(row["Ученое звание"]),
                    clean_text(row["Направление"]),
                    clean_text(row["Контакт"]),
                ),
            )

    return len(rows)


def export_assignments_to_excel(file_path, db_path=DB_PATH):
    init_db(db_path)
    rows = get_assignments_for_export(db_path)
    if not rows:
        raise ValueError("Нет назначений для выгрузки")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Назначения"

    headers = [
        "ФИО студента",
        "Группа",
        "Курс",
        "Тип работы",
        "Тема",
        "Руководитель",
        "Статус",
        "Дата изменения",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row["student_name"],
                row["study_group"],
                row["course"],
                row["work_type"],
                row["topic_title"],
                row["teacher_name"],
                STATUS_LABELS.get(row["status"], row["status"]),
                row["updated_at"],
            ]
        )

    format_assignments_sheet(sheet)
    workbook.save(file_path)


def format_assignments_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column, width in EXPORT_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index in (3,):
        column = get_column_letter(column_index)
        for cell in sheet[column]:
            cell.alignment = Alignment(horizontal="center", vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def validate_students(rows):
    seen_students = set()

    for index, row in enumerate(rows, start=3):
        full_name = clean_text(row["ФИО"])
        group = clean_text(row["Группа"])

        if not full_name:
            raise ValueError(f"Строка {index}: не заполнено ФИО студента")
        if not group:
            raise ValueError(f"Строка {index}: не заполнена группа")
        try:
            normalize_course(row["Курс"])
        except ValueError:
            raise ValueError(f"Строка {index}: курс должен быть 3 или 4")

        student_key = (full_name, group)
        if student_key in seen_students:
            raise ValueError(f"Строка {index}: студент повторяется в файле")
        seen_students.add(student_key)


def validate_teachers(rows):
    seen_teachers = set()

    for index, row in enumerate(rows, start=3):
        full_name = clean_text(row["ФИО"])
        position = clean_text(row["Должность"])

        if not full_name:
            raise ValueError(f"Строка {index}: не заполнено ФИО преподавателя")
        if not position:
            raise ValueError(f"Строка {index}: не заполнена должность")
        if full_name in seen_teachers:
            raise ValueError(f"Строка {index}: преподаватель повторяется в файле")
        seen_teachers.add(full_name)


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_course(value):
    course_text = clean_text(value)
    if not course_text:
        raise ValueError

    try:
        course_value = float(course_text)
    except ValueError:
        raise ValueError

    if not course_value.is_integer():
        raise ValueError

    course = int(course_value)
    if course not in (3, 4):
        raise ValueError

    return course
